import json, string
import xlsxwriter, xlwt, xlrd
from openpyxl import load_workbook
import os
alphabet = str(string.ascii_uppercase)
import expandUserPath

# workbook = xlrd.open_workbook(path)

def append(path, data):
    workbook = load_workbook(filename = expandUserPath.get(path))
    worksheet = workbook.worksheets[0]
    nextRow = worksheet.max_row + 1
    index = 1
    for item in data:
        worksheet.cell(
            row=nextRow,
            column=index,
            value=item)
        index += 1
    workbook.save(filename = expandUserPath.get(path));

# props
# - fileName
# - culloms
def write(props):
    # Make excell file
    workbook = xlsxwriter.Workbook(props['fileName'])
    worksheet = workbook.add_worksheet()

    # Set Styles
    label_format = workbook.add_format({
        'bold': True,
        'font_color': '#ffffff',
        'bg_color': '#392061'
    })

    # Write Titles
    index = 0
    for letter in alphabet:
        if letter in props['columns']:
            worksheet.write( letter+'1', props['columns'][letter]['label'].upper(), label_format)
            worksheet.set_column(index,index,props['columns'][letter]['width']) #
            index += 1
        else:
            break

    # Yes in another for loop because I want to keep functionality separate
    rownumber = 2
    for item in props['items']:
        for letter in alphabet:
            if letter in props['columns']:
                # key exists in item
                if props['columns'][letter]['label'] in item:
                    worksheet.write(letter+str(rownumber), item[props['columns'][letter]['label']] )
                # Write Blank
                else:
                    worksheet.write(letter+str(rownumber), ' ')
            else:
                break
        rownumber += 1

    workbook.close()














#
